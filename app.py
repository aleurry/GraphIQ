import streamlit as st
import pandas as pd
import plotly.express as px
import io
import numpy as np
from scipy import stats
from io import BytesIO
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_data(df, missing_threshold=0.5, zscore_threshold=3, numeric_range=None):
    report = []

    # Step 1: Drop columns or rows with excessive missing values
    missing_col_percent = df.isnull().mean()
    removed_cols = missing_col_percent[missing_col_percent >= missing_threshold].index.tolist()
    if removed_cols:
        report.append(f"Dropped columns with missing values above {missing_threshold * 100}%: {removed_cols}")
    df = df.loc[:, missing_col_percent < missing_threshold]

    missing_row_percent = df.isnull().mean(axis=1)
    rows_before = df.shape[0]
    df = df.loc[missing_row_percent < missing_threshold]
    rows_after = df.shape[0]
    report.append(f"Removed rows with missing values above {missing_threshold * 100}%. Rows before: {rows_before}, Rows after: {rows_after}")

    if "Name" in df.columns:
        name_missing = df["Name"].isna().sum() + (df["Name"] == "").sum()
        df = df[df["Name"].notna()]
        df = df[df["Name"] != ""]
        report.append(f"Removed {name_missing} rows where 'Name' was empty or missing.")

    # Step 2: Handling Text in Number Columns
    numeric_cols = df.select_dtypes(include=["number"]).columns
    for col in df.columns.difference(numeric_cols):
        if df[col].astype(str).str.isnumeric().mean() > 0.5:
            df[col] = pd.to_numeric(df[col], errors='coerce')
            report.append(f"Converted mostly numeric column '{col}' to numeric.")

    # Step 3: Handling missing values
    for col in df.select_dtypes(include=["number"]).columns:
        mean_value = df[col].mean()
        rounded_mean = np.ceil(mean_value)
        missing_count = df[col].isna().sum()
        df[col].fillna(rounded_mean, inplace=True)
        if missing_count > 0:
            report.append(f"Filled {missing_count} missing values in numeric column '{col}' with rounded mean ({rounded_mean}).")

    for col in ["Name", "name"]:
        if col in df.columns:
            # Step 1: Replace missing or empty values with 'Unknown'
            blank_count = df[col].isna().sum() + (df[col] == "").sum()
            df[col] = df[col].fillna("Unknown").replace("", "Unknown")

            # Step 2: Count and replace invalid names (those containing digits)
            digit_count = df[col].apply(lambda x: any(char.isdigit() for char in x)).sum()
            df[col] = df[col].apply(lambda x: "Unknown" if any(char.isdigit() for char in x) else x)
            
            # Step 3: Report the changes
            if blank_count > 0:
                report.append(f"Replaced {blank_count} missing or empty names in column '{col}' with 'Unknown'.")
            if digit_count > 0:
                report.append(f"Replaced {digit_count} invalid names containing numbers in column '{col}' with 'Unknown'.")
            if blank_count == 0 and digit_count == 0:
                report.append(f"No invalid or missing names found in column '{col}'.")


    # Fill missing values for categorical columns
    for col in df.select_dtypes(include=["object", "category"]).columns:
        if col != "Name":
            mode_value = df[col].mode()[0]
            missing_count = df[col].isna().sum()
            df[col].fillna(mode_value, inplace=True)
            if missing_count > 0:
                report.append(f"Filled {missing_count} missing values in categorical column '{col}' with mode ({mode_value}).")

    # Step 4: Outlier Detection and Handling
    for col in numeric_cols:
        z_scores = np.abs(stats.zscore(df[col]))
        outliers = (z_scores > zscore_threshold).sum()
        if outliers > 0:
            df[col] = np.where(
                z_scores > zscore_threshold,
                np.where(df[col] > df[col].mean(), df[col].quantile(0.95), df[col].quantile(0.05)),
                df[col]
            )
            report.append(f"Handled {outliers} outliers in column '{col}' using z-score threshold ({zscore_threshold}).")

    # Step 5: Handle out-of-range values
    if numeric_range:
        for col, (min_val, max_val) in numeric_range.items():
            if col in df.columns:
                out_of_range = ((df[col] < min_val) | (df[col] > max_val)).sum()
                df[col] = np.clip(df[col], min_val, max_val)
                report.append(f"Clipped {out_of_range} values in column '{col}' to range ({min_val}, {max_val}).")

    # Step 6: Handle duplicate rows
    duplicates = df.duplicated().sum()
    if duplicates > 0:
        df = df.drop_duplicates()
        report.append(f"Removed {duplicates} duplicate rows.")

    return df, report

def generate_chart_description(fig):
    if fig:
        chart_type = fig.layout.title.text if hasattr(fig.layout, 'title') else 'Visualization'

        # Default axis titles if they are not available
        x_axis_title = fig.layout.xaxis.title.text if hasattr(fig.layout, 'xaxis') and hasattr(fig.layout.xaxis, 'title') else 'X-Axis'
        y_axis_title = fig.layout.yaxis.title.text if hasattr(fig.layout, 'yaxis') and hasattr(fig.layout.yaxis, 'title') else 'Y-Axis'

        # Handle Line Graphs
        if chart_type == "Line Graph":
            description = (
                f"This is a Line Graph that shows the trend of {y_axis_title} over {x_axis_title}. "
                f"The data points are connected by lines to indicate changes in {y_axis_title} as {x_axis_title} increases. "
                f"The graph helps identify patterns, such as trends, fluctuations, and correlations between the two variables."
            )

        # Handle Bar Graphs
        elif chart_type == "Column Graph":
            description = (
                f"This is a Column Graph comparing the values of {y_axis_title} across different categories of {x_axis_title}. "
                f"Each bar represents a different category on the {x_axis_title} axis, and the height of the bar represents the value of {y_axis_title}. "
                f"This chart is useful for comparing quantities across discrete categories."
            )

        # Handle Heatmaps
        elif chart_type == "Heatmap":
            description = (
                f"This is a Heatmap showing the correlation between different numeric variables. "
                f"The colors in the heatmap represent the strength of correlation: darker colors indicate stronger correlations, "
                f"while lighter colors represent weaker correlations. The X and Y axes represent different variables."
            )

        # Handle Radial Charts
        elif chart_type == "Radial Chart":
            description = (
                f"This is a Radial Chart displaying the relationship between {x_axis_title} and {y_axis_title} in a circular format. "
                f"The chart uses polar coordinates where {x_axis_title} is represented by the angle and {y_axis_title} by the radius. "
                f"Radial charts are often used to show cyclic or periodic trends."
            )

        # Handle Funnel Charts
        elif chart_type == "Funnel Chart":
            description = (
                f"This is a Funnel Chart showing the progressive stages in a process. "
                f"The chart is typically used to visualize the drop-off or conversion rate between each stage of a process. "
                f"The {x_axis_title} represents the stages, and the {y_axis_title} shows the quantity at each stage."
            )

        # General case: When chart type is not specifically identified
        else:
            description = (
                f"This is a {chart_type}. It visualizes the relationship between {y_axis_title} and {x_axis_title}. "
                f"The chart allows us to understand how changes in {x_axis_title} affect {y_axis_title}. "
                f"Additional insights such as trends, patterns, or outliers can be observed from this visualization."
            )

        # For specific information, like ranges or unique data points
        if hasattr(fig.data[0], 'x') and hasattr(fig.data[0], 'y'):
            x_min, x_max = min(fig.data[0].x), max(fig.data[0].x)
            y_min, y_max = min(fig.data[0].y), max(fig.data[0].y)
            description += f" The range of {x_axis_title} is from {x_min} to {x_max}, and the range of {y_axis_title} is from {y_min} to {y_max}."

        return description
    return "No chart to describe."

def generate_report(df, chart, report, description):
    # Create an Excel report in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1. Write the cleaned data to the "Cleaned Data" sheet
        df.to_excel(writer, index=False, sheet_name="Cleaned Data")

        # 2. Add the explanation to the "Explanation" sheet
        workbook = writer.book
        worksheet = workbook.create_sheet("Explanation")

        # Combine data cleaning report with chart description
        combined_report = (
            ["Data Cleaning Report", "----------------------"] +
            report +
            ["", "Chart Description", "----------------"]
        )

        # Split chart description into multiple rows for readability
        wrapped_description = description.split(". ")
        combined_report += [line + "." for line in wrapped_description if line]

        # Write the combined report to the worksheet
        for idx, line in enumerate(combined_report):
            worksheet.cell(row=idx + 1, column=1, value=line)

        # Adjust column width and enable text wrapping
        column_width = max(len(line) for line in combined_report) + 5
        worksheet.column_dimensions['A'].width = min(column_width, 50)  # Limit max column width
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # 3. Add the chart to the "Chart" sheet
        # if chart is not None:
        #     # Convert chart to image (PNG format)
        #     chart_image = chart.to_image(format="png")
        #     chart_image_stream = io.BytesIO(chart_image)

        #     # Create a new sheet for the chart
        #     chart_sheet = workbook.create_sheet("Chart")
        #     # Add the image to the new sheet
        #     img = openpyxl.drawing.image.Image(chart_image_stream)
        #     chart_sheet.add_image(img, "A1")
    
    # Save the file to memory and return the bytes
    output.seek(0)
    return output

def load_css():
    """Load CSS for the Streamlit app from an external file."""
    with open('style.css') as f:
        st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)

def main():
    st.set_page_config(page_title="Data Cleaning and Visualization", layout="wide")
    
    # Load custom CSS
    load_css()

    st.image("images/logo.png")
    # Welcome page with a "Get Started" button
    with st.container():
        if 'started' not in st.session_state:
            st.session_state['started'] = False
        if 'visualization_started' not in st.session_state:
            st.session_state['visualization_started'] = False

        if not st.session_state.started:
            st.markdown('<p class="big-font">Welcome to GraphIQ!</p>', unsafe_allow_html=True)
            st.markdown('<p class="medium-font">We are excited to have you on board. With GraphIQ, you can easily clean, analyze, and visualize your data like never before. <br> Get started today and unlock powerful insights with just a few clicks!</p>', unsafe_allow_html=True)
            if st.button("Get Started"):
                st.session_state.started = True

        if st.session_state.started:
            st.markdown('<p class="steps-font">Step 1: Upload CSV File</p>', unsafe_allow_html=True)

            # Step 1: File Upload
            uploaded_file = st.file_uploader("Upload a CSV File", type="csv")
            
            if uploaded_file is not None:
                df = pd.read_csv(uploaded_file)
                st.write("Uploaded CSV:")
                st.dataframe(df)

                # Automatically clean data when uploaded
                df, report = clean_data(df)
                st.write("Cleaned Data:")
                st.dataframe(df)
                
                # Display the cleaning process report
                st.subheader("Data Cleaning Report")
                for item in report:
                    st.write("- " + item)

                st.markdown('<p class="steps-font">Step 2: Proceed to Visualization</p>', unsafe_allow_html=True)
                # Add a button to start visualization
                if st.button("Start Visualization"):
                    st.session_state.visualization_started = True

                if st.session_state.visualization_started:
                    # Visualization and report generation
                    visualize_and_report(df, report)
                

def visualize_and_report(df, report):
    st.subheader("Choose Visualization Type")
    
    # Determine available chart types based on the data
    numeric_columns = df.select_dtypes(include=np.number).columns.tolist()
    categorical_columns = df.select_dtypes(include=['object', 'category']).columns.tolist()
    available_charts = []

    if len(numeric_columns) > 1:
        available_charts.append("Line Graph")
        available_charts.append("Heatmap")
    if len(categorical_columns) > 0 and len(numeric_columns) > 0:
        available_charts.append("Column Graph")
        available_charts.append("Funnel Chart")
    if len(numeric_columns) > 0 and len(categorical_columns) > 0:
        available_charts.append("Radial Chart")

    if not available_charts:
        st.warning("No suitable charts available for the current dataset.")
        return

    visualization_type = st.selectbox(
        "Select a chart type:",
        available_charts
    )

    st.subheader("Select Columns for Visualization")
    if visualization_type in ["Line Graph", "Column Graph", "Funnel Chart", "Radial Chart"]:
        x_axis = st.selectbox("X-Axis", options=categorical_columns if visualization_type in ["Column Graph", "Funnel Chart", "Radial Chart"] else numeric_columns)
        y_axis = st.selectbox("Y-Axis", options=numeric_columns)
    elif visualization_type == "Heatmap":
        st.info("Heatmap will use all numeric columns for correlation matrix.")

    fig = None

    try:
        if visualization_type == "Line Graph":
            fig = px.line(df, x=x_axis, y=y_axis, title="Line Graph")
        elif visualization_type == "Column Graph":
            fig = px.bar(df, x=x_axis, y=y_axis, title="Column Graph")
        elif visualization_type == "Heatmap":
            fig = px.imshow(df[numeric_columns].corr(), title="Heatmap")
        elif visualization_type == "Radial Chart":
            # Check if data is applicable for radial chart
            if x_axis not in categorical_columns or y_axis not in numeric_columns:
                st.warning("Radial chart requires a categorical column for the X-axis and a numeric column for the Y-axis.")
            elif df[x_axis].nunique() < 2:
                st.warning("Radial chart requires at least two unique values in the X-axis column.")
            else:
                try:
                    fig = px.line_polar(df, r=y_axis, theta=x_axis, line_close=True, title="Radial Chart")
                except Exception as e:
                    st.warning("Radial chart cannot be generated with the selected data.")

        elif visualization_type == "Funnel Chart":
            fig = px.funnel(df, x=x_axis, y=y_axis, title="Funnel Chart")

        if fig:
            st.plotly_chart(fig)

            # Auto-generate a description for the chart
            description = generate_chart_description(fig)
            st.subheader("Chart Description")
            st.write(description)

        # Step 4: Report Download
        st.markdown('<p class="steps-font">Step 3: Generate a Copy of Cleaned Data with Visualization and Reports</p>', unsafe_allow_html=True)
        if st.button("Generate Reports"):
            report_file = generate_report(df, fig, report, description)
            st.markdown('<p class="steps-font">Step 4: You Can Now Save a Copy:)</p>', unsafe_allow_html=True)
            st.download_button(
                label="Download",
                data=report_file,
                file_name="Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"An error occurred while generating the {visualization_type}: {e}")


if __name__ == "__main__":
    main()
